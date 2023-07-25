import os
import PyPDF2 
from openpyxl import Workbook
import tkinter as tk
from tkinter.filedialog import askdirectory 
from tkinter import filedialog, messagebox





def obtener_numero_paginas_pdf(archivo_pdf):
    try:
        with open(archivo_pdf, 'rb') as archivo:
            lector_pdf = PyPDF2.PdfReader(archivo)
            numero_paginas = len(lector_pdf.pages) 
            ruta_e = 0, 
            exc = 0       
            return numero_paginas, ruta_e, exc
    except Exception as e:
        print('XXXXXXXXXXXXXXXXXXXXX')        
        ruta_e = archivo_pdf
        numero_paginas = 0
        exc = str(type(e).__name__)
        return numero_paginas, ruta_e, exc
        
    


    
def guardar_archivo_excel(*args, final=False, libro_excel=None):
    
    # Inicializar el libro de Excel y la hoja de cálculo    
    if libro_excel is None:
        libro_excel = Workbook()
        

    if final:
        hoja_resumen = libro_excel.create_sheet(title="Resumen")  
        hoja_resumen.cell(row=1, column=1, value="Total de Expedientes_81")
        hoja_resumen.cell(row=1, column=2, value="Total de Expedientes_54")
        hoja_resumen.cell(row=1, column=3, value="Total de Folios")                 
                                                                                                                                                                                                                                                                                                                                                                                                                                                                        
        hoja_resumen.cell(row=2, column=1, value=args[0])
        hoja_resumen.cell(row=2, column=2, value=args[1])
        hoja_resumen.cell(row=2, column=3, value=args[2])                    

    elif len(args) == 3:
        
        hoja_calculo = libro_excel.active
        hoja_calculo.title = "Folios X expedientes"
        if hoja_calculo.max_row == 1:
        # Encabezados de la hoja de cálculo
            hoja_calculo.cell(row=1, column=1, value="Expedientes")
            hoja_calculo.cell(row=1, column=2, value="Folios")
            
        hoja_calculo.cell(row=args[0], column=1, value=args[1])
        hoja_calculo.cell(row=args[0], column=2, value=args[2]) 

    elif len(args) == 1:             
        if "Errors" not in libro_excel.sheetnames:
            hoja_errors = libro_excel.create_sheet(title="Errors")
            hoja_errors.cell(row=1, column=1, value="Ruta")  
            hoja_errors.cell(row=1, column=2, value="Tipo de Problema")                      
        else:
            hoja_errors = libro_excel.get_sheet_by_name("Errors")
            
        hoja_errors.cell(row=args[0]['fila'], column=1, value=args[0]['ruta_e'])
        hoja_errors.cell(row=args[0]['fila'], column=2, value=args[0]['error'])
                                  

    return libro_excel 




def buscar_archivos_pag(ruta):
    print("Procesando directorio:", ruta)
    total_paginas_pdf = 0    
    problemas = {}
    for subdirectorio, _, archivos_subdir in os.walk(ruta):#se ingresa al subdirectorio 54 o 81        
        for archivo in archivos_subdir:# muestra archivo

            if archivo.endswith('.pdf'):
                archivo_completo = os.path.join(subdirectorio, archivo)  
                cantidad_pag, ruta_e, exc = obtener_numero_paginas_pdf( archivo_completo)
                if ruta_e != 0 and exc != 0:    
                    problemas['ruta_e'] = ruta_e
                    problemas['error'] = exc
                                                
                total_paginas_pdf = total_paginas_pdf + cantidad_pag                    
                
    return total_paginas_pdf, problemas




def contar_dir_y_pag(ruta): 
    print("Iniciando proceso...") 
    total_directorios_81 = 0
    total_directorios_54 = 0   
    total_paginas_pdf = 0    
    fila_actual = 2
    fila = 2
    wb = guardar_archivo_excel(final=False)
     
    for directorio_actual, directorios, archivos in os.walk(ruta):        
        for directorio in directorios:
            if directorio.startswith('81') or directorio.startswith('54'):                 
                directorio_completo = os.path.join(directorio_actual, directorio)                
                folios, problemas = buscar_archivos_pag(directorio_completo)                
                wb = guardar_archivo_excel(fila_actual, directorio, folios, libro_excel=wb)
                if len(problemas) == 2:
                   problemas['fila'] = fila
                   wb = guardar_archivo_excel(problemas, libro_excel=wb)                   
                   fila += 1 
                total_paginas_pdf = total_paginas_pdf + folios                              
                fila_actual += 1 
                if directorio.startswith('81'):
                    total_directorios_81 += 1
                else:
                    total_directorios_54 +=  1 
        
    
    wb = guardar_archivo_excel(total_directorios_81, total_directorios_54, total_paginas_pdf, final=True, libro_excel=wb)
    wb.save('informacion.xlsx')
    
    
    
    

def seleccionar_ruta():
    
    
    ruta = filedialog.askdirectory()
    if ruta:       
        contar_dir_y_pag(ruta)        
        messagebox.showwarning("Advertencia", "El proceso finalizó con éxito")          
    else:
        messagebox.showwarning("Advertencia", "No se seleccionó ninguna ruta.")




# Crear ventana principal
ventana = tk.Tk()
ventana.title("Contador de Directorios y Páginas en PDF")
ventana.geometry("400x200")

# Etiqueta y botón para seleccionar ruta
etiqueta = tk.Label(ventana, text="Seleccione una ruta de directorio:")
etiqueta.pack(pady=10)

boton_ruta = tk.Button(ventana, text="Seleccionar ruta", command=seleccionar_ruta)
boton_ruta.pack(pady=5)

# Ejecutar ventana
ventana.mainloop()